import mysql.connector
import numpy as np
import pandas as pd
import pymysql
import sqlalchemy
from os import listdir
import os
from typing import List
import xlsxwriter
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.workbook.workbook import save_workbook
from openpyxl.utils import get_column_letter
import getpass



class ExcelFilesDetails():
    def __init__(self, input_path: str, suffix: str):
        self._input_path = input_path
        self._suffix = suffix
        self.mls_whole_name = self.long_excel_filenames()
        self.mls_short_names = self.short_excel_filenames()
        self.mls_tab_names = self.get_tab_excel_list()
        self.mdic_files_and_tabs = self.create_dictionary()

    def long_excel_filenames(self) -> List[str]:
        return list(filter(lambda x: self._suffix in x, os.listdir(self._input_path)))

    def short_excel_filenames(self) -> List[str]:
        ls_whole_names = list(filter(lambda x: self._suffix in x, os.listdir(self._input_path)))
        ls_short_names = []
        for i in range(len(ls_whole_names)):
            s_temp = ls_whole_names[i]
            s_ticker = s_temp[:-4]
            ls_short_names.append(s_ticker)
        return ls_short_names

    def get_tab_excel_list(self) -> List[str]:
        tab_list = []
        for i in range(len(self.mls_whole_name)):
            xlsx = pd.ExcelFile(self.mls_whole_name[i])
            ls_file_sheets = xlsx.sheet_names
            tab_list.append(ls_file_sheets)
        return tab_list

    def create_dictionary(self) -> dict:
        files_names_and_tabs = dict(zip(self.mls_short_names, self.mls_tab_names))
        return files_names_and_tabs

    def dataFrameSumary(self) -> pd.DataFrame():
        return pd.DataFrame.from_dict(self.create_dictionary(), orient='index')




class CreateDataFrame:
    def __init__(self,path, file_name: str, sheet_name=None, set_index=None):
        self._path=path
        self._file_name = file_name
        self._sheet_name = sheet_name
        self._set_index = set_index
        self.mdf = self.create_data_frame_from_excel()

    def create_data_frame_from_excel(self):
        os.chdir(self._path)
        listOfDfs = []
        if self._sheet_name == None:
            dfsFromExcel = pd.read_excel(self._file_name, sheet_name=None)
            lTabs = list(dfsFromExcel)
            for k in range(len(lTabs)):
                temp = pd.read_excel(self._file_name, sheet_name=lTabs[k])
                listOfDfs.append(temp)
            di = dict(zip(lTabs, listOfDfs))
            return di
        else:
            df_from_excel = pd.read_excel(self._file_name, sheet_name=self._sheet_name)
            if self._set_index != None:
                df_from_excel = pd.read_excel(self._file_name, sheet_name=self._sheet_name)
                return df_from_excel.set_index(self._set_index)
            else:
                return df_from_excel

    def get_columns(self) -> List[str]:
        llsColumns = []
        if type(self.mdf) is dict:
            for i in self.mdf.keys():
                temp = list(self.mdf[i].columns)
                llsColumns.append(temp)
        else:
            return self.mdf.columns
        return llsColumns

    def get_dimension(self):
        ltShapes = []  # list of touples
        if type(self.mdf) is dict:
            for i in self.mdf.keys():
                temp = list(self.mdf[i].shape)
                ltShapes.append(temp)
        else:
            return self.mdf.shape

        return ltShapes


    def modify_columns_data_frame(self, columns_name: str,
                                  l_fill_in: str) -> pd.DataFrame:  # TODO expand possibility to remove columns, the same with rows
        base_data_frame = self.mdf
        new_col = [l_fill_in] * len(base_data_frame)
        base_data_frame[columns_name] = new_col
        modified_df = base_data_frame
        return modified_df

class OutputInExcel:
    def __init__(self, FileName, Path,SheetNames=None):
        self._sFileName = FileName
        self._lsSheetName = SheetNames
        self._sPath = Path

    def createResultsToPresent(self, ldfToSave, formatStyle=None, colRange=None) -> None:
        if not os.path.isdir(self._sPath):
            os.makedirs(self._sPath)
            os.chdir(self._sPath)
        if len(self._lsSheetName) == 1 and type(ldfToSave) == pd.DataFrame():

            oDataToExcel = pd.ExcelWriter(self._sFileName, engine='xlsxwriter')
            ldfToSave.to_excel(oDataToExcel, sheet_name=self._lsSheetName[0])
            workbook = oDataToExcel.book
            worksheet = oDataToExcel.sheets[self._lsSheetName[0]]
            if (formatStyle == 'percentage'):
                fixedFormat = workbook.add_format({'num_format': '0%'})
                worksheet.set_column(colRange, None, fixedFormat)
            oDataToExcel.save()
            if (formatStyle == 'CommaFormat'):
                fixedFormat = workbook.add_format({'num_format': '#,##0.00%'})
                worksheet.set_column(colRange, None, fixedFormat)
            oDataToExcel.save()
            oDataToExcel.close()
        else:
            if not os.path.isdir(self._sPath):
                os.makedirs(self._sPath)
                oDataToExcel = pd.ExcelWriter(self._sFileName, engine='xlsxwriter')
                for i in range(len(ldfToSave)):
                    ldfToSave[i].to_excel(oDataToExcel, sheet_name=self._lsSheetName[i])
                    oDataToExcel.save()
                    oDataToExcel.close()

    def appendDfToExisingExel(self,filename,fileLocation,df,sheet_name,startrow=None,truncate_sheet=False,startcol=None):

        """Append a DataFrame [df] to existing Excel file [filename] into Sheet[sheet_name]
               If [filename] does not exist then this function will create it.

            Returns: None

            ."""
        os.chdir(fileLocation)
        writer=pd.ExcelWriter(filename,engine='openpyxl')
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError=IOError
        try:
            #try to open an existing workbook
            writer.book=load_workbook(filename)
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow=writer.book.sheetnames.max_row
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet

                idx=writer.book.sheetnames.index(sheet_name)
                #remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                #create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name,idx)
            writer.sheets={ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            #file does not exist yet, we will create it
            pass
        if startrow is None:
            startrow=0
        df.to_excel(writer,sheet_name,startrow=startrow,startcol=startcol,header=False,index=False)
        ws=writer.book['INPUT_3M']

        cell=ws.cell(column=6,row=2)
        cell=5

        writer.save()

    def flexibleInsertingScalar(self,cell_row,cell_col,value,tab_name):

            writer = pd.ExcelWriter(self._sFileName, engine='openpyxl')
            writer.book = load_workbook(self._sFileName)
            ws = writer.book[tab_name]

            cell=ws.cell(column=cell_col,row=cell_row)
            cell.value=value
            writer.save()

    def insertRange(self,tab_name,iterativeObj,colIndicator,startrow):
        writer = pd.ExcelWriter(self._sFileName, engine='openpyxl')
        writer.book = load_workbook(self._sFileName)
        ws = writer.book[tab_name]
        for i in range(startrow,len(iterativeObj)+2):
            cell = ws.cell(column=colIndicator, row=i)
            cell.value=iterativeObj[i-2][0]
        writer.save()

    def insertPngFile(self,tab_name,image_name,anchore):
        writer = pd.ExcelWriter(self._sFileName, engine='openpyxl')
        writer.book = load_workbook(self._sFileName)

        ws = writer.book[tab_name]
        img=Image(image_name+'.png')
        ws.add_image(img,anchore)
        writer.save()







if __name__ == "__main__":
    pass
